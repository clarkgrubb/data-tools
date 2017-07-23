#!/usr/bin/env python3

import argparse
import csv
import re
import sys

import openpyxl

REGEX_CSV_SUFFIX = re.compile(r'.csv$', re.I)  # pylint: disable=no-member
REGEX_XLSX_SUFFIX = re.compile(r'.xlsx$', re.I)  # pylint: disable=no-member
REGEX_INVALID_SHEETNAME_CHARS = re.compile(r'[][*?/\.]')
REGEX_SPACES = re.compile(' +')
MAX_SHEETNAME_LENGTH = 31
ENCODING = 'utf-8'
START_INDEX = 0 if openpyxl.__version__.startswith('1.') else 1


def path_to_sheetname(path):
    sheetname = REGEX_CSV_SUFFIX.sub('', path)
    sheetname = REGEX_INVALID_SHEETNAME_CHARS.sub(' ', sheetname)
    sheetname = REGEX_SPACES.sub(' ', sheetname)

    return sheetname[0:MAX_SHEETNAME_LENGTH].strip()


def csv_to_xlsx(input_files, output_file):
    wb = openpyxl.Workbook()
    sheetnames = {}

    for filenum, input_file in enumerate(input_files):
        with open(input_file, encoding=ENCODING) as f:
            rows = csv.reader(f, dialect=csv.excel)
            if filenum == 0:
                ws = wb.get_active_sheet()
            else:
                ws = wb.create_sheet()
            sheetname = path_to_sheetname(input_file)
            if sheetname in sheetnames:
                raise ValueError('files {} and {} result in the same sheet '
                                 'name: "{}"'.format(input_file,
                                                     sheetnames[sheetname],
                                                     sheetname))
            sheetnames[sheetname] = input_file
            ws.title = sheetname
            for rownum, row in enumerate(rows, start=START_INDEX):
                for colnum, value in enumerate(row, start=START_INDEX):
                    # WHAT ABOUT DATES
                    ws.cell(row=rownum, column=colnum).value = value   # pylint: disable=no-member

    wb.save(output_file)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('input_files',
                        nargs='+',
                        metavar='CSV_FILE')
    parser.add_argument('--output-file', '-o',
                        dest='output_file',
                        required=True)

    args = parser.parse_args()

    if not REGEX_XLSX_SUFFIX.search(args.output_file):
        sys.stderr.write('ERROR: output file must have .xlsx '
                         'suffix: {}\n'.format(args.output_file))
        sys.exit(1)

    csv_to_xlsx(args.input_files, args.output_file)


if __name__ == '__main__':
    main()
